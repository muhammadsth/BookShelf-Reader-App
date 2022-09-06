from PIL import Image
from pytesseract import image_to_string, image_to_data, image_to_boxes
from ArabicOcr import arabicocr
import cv2
from openpyxl import load_workbook, Workbook, drawing
#OS folder for photos open_folder
#Edge detection opencv or any other edge detection algrithms (laplace?)
#Web scraping
wb = load_workbook('Sample.xlsx')
#wb = load_workbook('جرد كتب بابا .xlsx')
ws = wb.active

def do_pairs(k):
    result = []
    for i in range(len(k)):
        if(i == len(k) - 2 and len(k) % 2 != 0):
            result.append([k[len(k) - 1]])
        if(i == len(k) - 1):
            break
        if(i % 2 == 0):
            result.append([k[i], k[i+1]])
            i+=1
    return result



for i in range(364, 427):
    im = 'IMG_0'+ str(i)+'.jpg'
    outim = 'out_'+im
    try:
        results  = arabicocr.arabic_ocr(im, outim)
        ws.append(['_____________________________________________________________', im, '_____________________________________________________________'])

        words = []
        for i in range(len(results)):
            word = results[i][1]
            words.append(word)

        results = do_pairs(words)
    
        for j in results:
            ws.append(j)
            wb.save('Sample.xlsx')
    
        wb.save('Sample.xlsx')
    except:
        print('Move on')
        i+=1



    


    #img = drawing.image.Image(outim)
    #ws.add_image(outim)

        


"""
with open ('file.txt','w',encoding='utf-8')as myfile:
		myfile.write(str(words))
        
"""

#kim = Image.open(im)
#text = image_to_string(kim, lang='eng')


""" 
#img = drawing.image.Image('out_IMG-1037.jpg')
#width, height = img.size
#print(width, height)
#img.resize()
#width, height = img.size
#print(width, height)
#img.anchor = 'A1'
#ws.add_image(img)
#wb.save('Sample4.xlsx')


"""